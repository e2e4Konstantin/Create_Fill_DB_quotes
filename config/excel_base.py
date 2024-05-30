import openpyxl
import os


class ExcelBase:
    """Базовый класс для операций с Excel файлом"""

    def __init__(self, filepath: str = None):
        self.filepath = filepath
        self.workbook = None
        self.worksheet = None
        self.max_row = 0
        self.max_column = 0


    def __enter__(self):
        self.open_file()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close_file()

    def __str__(self):
        return (
            f"Excel file: {self.filepath}, worksheet: "
            f"{self.worksheet.title}"
        )

    def is_file_in_use(self) -> bool:
        """Проверка что файл используется другой программой"""
        return os.path.isfile(self.filepath) and not os.access(
            self.filepath, os.W_OK
        )

    def open_file(self):
        """Открывает файл или создает его"""
        if self.is_file_in_use():
            raise IOError(f"{self.filepath} файл используется другой программой.")
        try:
            self.workbook = openpyxl.load_workbook(self.filepath)
        except IOError:
            self.workbook = openpyxl.Workbook()

    def close_file(self):
        if self.workbook:
            self.workbook.close()

    def save_file(self):
        """Сохраняет файл"""
        if self.is_file_in_use():
            self.close_file()
            raise IOError(f"{self.filepath} файл используется другой программой")
        else:
            if self.workbook:
                self.workbook.save(self.filepath)

