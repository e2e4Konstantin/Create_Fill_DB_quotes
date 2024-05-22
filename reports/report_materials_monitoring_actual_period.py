import sqlite3
from icecream import ic
import numpy as np
from openpyxl.utils import get_column_letter
from typing import Optional

from config import dbTolls, LocalData, ROUNDING
from reports.sql_materials_report import sql_materials_reports
from collections import namedtuple
from dataclasses import dataclass, field

from reports.report_excel_config import ExcelReport
from files_features import output_message_exit, output_message


PriceHistory = namedtuple(
    typename="PriceHistory", field_names=["history_index", "history_price"]
)
PriceHistory.__annotations__ = {"history_index": int, "history_price": float}

def sum_of_differences(values):
    if values is None or len(values) <= 3:
        return 1
    mean = sum(values) / len(values)
    return sum(abs(value - mean) for value in values)


def abbe_criterion(signal):
    """Критерий Аббе."""
    if signal is None or len(signal) <= 3:
        return 0
    differences = np.diff(signal) ** 2
    squares = np.sum(differences)
    return np.sqrt(squares / (len(signal) - 1))



@dataclass
class Material:
    code: str
    name: str
    base_price: float
    index_num: int
    actual_price: float
    monitoring_index: int
    monitoring_price: float
    transport_flag: bool = False
    transport_code: str = ""
    transport_base_price: float = 0.0
    transport_numeric_ratio: float = 0.0
    transport_actual_price: float = 0.0
    gross_weight: float = 0.0
    history: list[PriceHistory] = field(default_factory=list)
    len_history: int = 0
    abbe_criterion: float = 0.0
    mean_history: float = 0.0
    std_history: float = 0.0

    def __post_init__(self):
        if self.history is None:
            raise ValueError("history is None")
        self.len_history = len(self.history)
        self.abbe_criterion = round(
            abbe_criterion([x.history_price for x in self.history]), 3
        )
        self.mean_history = np.mean([x.history_price for x in self.history])
        self.std_history = np.std([x.history_price for x in self.history])
        self.transport_numeric_ratio = round(
            self.transport_actual_price / self.transport_base_price, ROUNDING
        ) if self.transport_base_price != 0 else 0




def _get_material_price_history(
    db: dbTolls, transport_cost_id: int, history_depth: int
) -> list[PriceHistory]:
    """Получить историю цен на материал по id."""
    history = db.go_select(
        sql_materials_reports[
            "select_historical_prices_for_materials_id_not_empty_actual_price"
        ],
        (transport_cost_id, history_depth),
    )
    result = [
        PriceHistory(history_index=row["index_num"], history_price=row["actual_price"])
        for row in history
    ]
    result.sort(reverse=False, key=lambda x: x.history_index)
    return result


def _materials_constructor(
    db: dbTolls, row: sqlite3.Row, history_depth: int
) -> Material:
    material = Material(
        code=row["code"],
        name=row["title"],
        base_price=row["base_price"],
        index_num=row["index_num"],
        actual_price=row["actual_price"],
        monitoring_index=row["monitoring_index_num"],
        monitoring_price=row["monitoring_price"],
        transport_flag=bool(row["transport_flag"]),
        transport_code=row["transport_code"],
        transport_base_price=row["transport_base_price"],
        transport_actual_price=row["transport_actual_price"],
        gross_weight=row["gross_weight"],
        history=_get_material_price_history(db, row["ID_tblMaterial"], history_depth),
    )
    return material


def _get_materials_with_monitoring( db_file: str, history_depth: int) -> list[Material] | None:
    """Стоимость материалов с данными последнего загруженного мониторинга."""
    table = None
    with dbTolls(db_file) as db:
        # получить id записей материалов с максимальным индексом периода и данными мониторинга
        materials = db.go_select(
            sql_materials_reports["select_records_for_max_index_with_monitoring"]
        )
        table = [_materials_constructor(db, material, history_depth) for material in materials]
    return table if table else None



def _header_create(table: list[Material], view_history_depth: int) -> str:
    header = [
        "No",
        "шифр",
        "название",
        "базовая стоимость",
        "actual index",
        "actual price",
        "monitoring index",
        "monitoring price",
        "transport flag",
        "transport code",
        "transport base price",
        "transport numeric ratio",
        "transport actual price",
        "gross weight",
    ]
    header_calculated = [
            "_",
            "стоимость перевозки",
            "цена для загрузки",
            "предыдущий индекс",
            "текущий индекс",
            "рост абс.",
            "рост %",
            "_",
            "abbe критерий",
            'абс.', 'флаг',
            'процент',
        ]
    history_sizes = set([material.len_history for material in table])
    ic(history_sizes)  #  {1, 4, 5}
    max_history_len = max(history_sizes)
    history_header = []
    #  найти материал с длинной историей и записать заголовок
    for material in table:
        if material.len_history == max_history_len:
            # ic(material)
            history_header = [x.history_index for x in material.history]
            #
            if max_history_len > view_history_depth:
                history_header = history_header[-view_history_depth:]
                max_history_len = len(history_header)
            break
    final_header = [*header[:4], *history_header, *header[4:], *header_calculated]
    return final_header, max_history_len


def _material_row_create(material: Material, row_number: int, max_history_len: int):
    base_value = [
        0,
        material.code,
        material.name,
        material.base_price,
        material.index_num,
        material.actual_price,
        material.monitoring_index,
        material.monitoring_price,
        material.transport_flag,
        material.transport_code,
        material.transport_base_price,
        material.transport_numeric_ratio,
        material.transport_actual_price,
        material.gross_weight,
        "",
    ]
    history_value = [x.history_price for x in material.history]
    if len(history_value) < max_history_len:
        for _ in range(max_history_len - len(history_value)):
            history_value.insert(0, None)
    else:
        history_value = history_value[-max_history_len:]
    #
    start_formula_column = len(base_value) + max_history_len

    cols = {
        "base_price": get_column_letter(start_formula_column - max_history_len - 11),
        "actual_price": get_column_letter(start_formula_column - 9),
        "monitoring_price": get_column_letter(start_formula_column - 7),
        "transport_flag": get_column_letter(start_formula_column - 6),
        "transport_base_price": get_column_letter(start_formula_column - 4),
        "transport_numeric_ratio": get_column_letter(start_formula_column - 3),
        "transport_actual_price": get_column_letter(start_formula_column - 2),
        "gross_weight": get_column_letter(start_formula_column -1),
        "transport_price": get_column_letter(start_formula_column + 1),
        "result_price": get_column_letter(start_formula_column + 2),
        "previous_index": get_column_letter(start_formula_column + 3),
        "result_index": get_column_letter(start_formula_column + 4),
        "abbe_criterion": get_column_letter(start_formula_column + 8),
        "absolute_price_change": get_column_letter(start_formula_column + 9),
    }
    # ic(start_formula_column)
    # ic(cols["transport_flag"])
    # ic(cols["transport_price"])
    # ic(cols["gross_weight"])

    formulas = [
        f"={cols['transport_base_price']}{row_number}*{cols['transport_numeric_ratio']}{row_number}*{cols['gross_weight']}{row_number}/1000",
        # f"={cols['transport_actual_price']}{row_number}*{cols['gross_weight']}{row_number}/1000",
        f"=ROUND(IF({cols['transport_flag']}{row_number}, ({cols['monitoring_price']}{row_number}-{cols['transport_price']}{row_number}), {cols['monitoring_price']}{row_number}),2)",
        f"={cols['actual_price']}{row_number}/{cols['base_price']}{row_number}",
        f"={cols['result_price']}{row_number}/{cols['base_price']}{row_number}",
        f"={cols['result_index']}{row_number}-{cols['previous_index']}{row_number}",
        f"=({cols['result_index']}{row_number}/{cols['previous_index']}{row_number})-1",
        "",
        material.abbe_criterion,
        f"=ROUND(ABS({cols['result_price']}{row_number}-{cols['actual_price']}{row_number}),3)",
        f'=IF({cols["absolute_price_change"]}{row_number}<=1.4*{cols["abbe_criterion"]}{row_number}, "", 1)',
        f"=({cols['result_price']}{row_number}/{cols['actual_price']}{row_number})-1",
    ]
    mod_row = [*base_value[:4], *history_value, *base_value[4:], *formulas]
    # mod_row.append(abbe)
    return mod_row

def _materials_monitoring_report_output(table):
    """Напечатать отчет по мониторингу материалов"""
    sheet_name = "materials"
    file_name = "report_monitoring.xlsx"
    with ExcelReport(file_name) as file:
        sheet = file.get_sheet(sheet_name)
        ic(sheet)
        #
        header, max_history_len = _header_create(table)
        file.write_header(sheet.title, header)
        #
        row = 2
        for i, material in enumerate(table):
            value_row = _material_row_create(material, row, max_history_len)
            value_row[0] = i + 1
            file.write_row(sheet_name, value_row, row)
            file.write_material_format(sheet_name, row, max_history_len)
            row += 1
        ic()

def _modern_materials_monitoring_report_output(
    table: list[Material], view_history_depth: int, sheet_name: str, file_name: str
) -> None:
    """Напечатать отчет по мониторингу материалов"""
    with ExcelReport(file_name) as file:
        sheet = file.get_sheet(sheet_name)
        ic(sheet)
        #
        header, max_history_len = _header_create(table, view_history_depth)
        file.write_header(sheet.title, header)
        #
        row = 2
        for i, material in enumerate(table):
            value_row = _material_row_create(material, row, max_history_len)
            value_row[0] = i + 1
            file.write_row(sheet_name, value_row, row)
            file.write_material_format(sheet_name, row, max_history_len)
            row += 1
        ic()


def _create_history_price_row(material: Material, max_history_len: int) -> list[Optional[float]]:
    """Создать строку с историей цен на материал"""
    first = [None] * (max_history_len - len(material.history))
    res = (
        first
        + [float(data.history_price) for data in material.history]
        + [
            round(material.abbe_criterion, 3)
            if material.abbe_criterion is not None
            else None
        ]
        + [
            round(material.mean_history, 3)
            if material.mean_history is not None
            else None
        ]
        + [
            round(material.std_history, 3)
            if material.std_history is not None
            else None
        ]
    )
    return res


def _material_price_history_report_output(
    table, sheet_name: str, file_name: str
) -> list[str] | None:
    """Напечатать отчет по ценам материалов"""
    max_history_len = max(material.len_history for material in table)
    for material in table:
        if material.len_history == max_history_len:
            history_header = [str(x.history_index) for x in material.history]
            break
    history_header = ["No", "code", *history_header, "abbe criterion", "mean", "std"]

    row = 2
    with ExcelReport(file_name) as file:
        sheet = file.get_sheet(sheet_name)
        if history_header:
            sheet.append(history_header)
        for i, material in enumerate(table):
            price_row = _create_history_price_row(material, max_history_len)
            price_row = [i + 1, material.code, *price_row]
            sheet.append(price_row)
            row += 1


def last_period_main():
    location = "office"  # office  # home
    local = LocalData(location)
    ic()

    table = _get_materials_with_monitoring(local.db_file, history_depth=10)
    ic(len(table))
    ic(table[5])
    # for material in table[:2]:
    #     ic(material)

    sheet_name = "materials"
    file_name = "report_monitoring.xlsx"
    _modern_materials_monitoring_report_output(
        table, view_history_depth=2, sheet_name =sheet_name, file_name = file_name
    )
    # _material_price_history_report_output(
    #     table, sheet_name="price materials history", file_name=file_name
    # )

if __name__ == "__main__":
    last_period_main()


