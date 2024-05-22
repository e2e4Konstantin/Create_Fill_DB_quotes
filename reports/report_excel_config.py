
from openpyxl.styles import Font, PatternFill, numbers, DEFAULT_FONT, Color
from config import ExcelBase
from openpyxl.utils import get_column_letter


class ExcelReport(ExcelBase):
    def __init__(self, filename: str = None):
        super().__init__(filename)
        self.sheet_names = [
            "transport",
            "materials",
        ]
        self.colors = {
            "transport": "00FF9900",
            "materials": "0099CC00",
            "header": "00E4DFEC",
            "default": "0099CCFF",
            "calculate": "00F7F7F7",
        }
        self.fonts = {
            "default": Font(name="Arial", size=8, bold=False, italic=False),
            "default_bold": Font(name="Arial", size=8, bold=True, italic=False),
            "green_bold": Font(name="Arial", size=8, bold=True, color="006600"),
            "grey": Font(name="Arial", size=8, bold=False, color=Color("808080")),
            "blue": Font(name="Arial", size=8, bold=False, color="1F497D"),
            "result_bold": Font(name="Arial", size=8, bold=True, color="990000"),
        }
        self.fills = {
            "calculate": PatternFill("solid", fgColor=self.colors["calculate"]),
            "header": PatternFill("solid", fgColor=self.colors["header"]),
        }
        self.number_format = "# ##0.00"

    #
    def __enter__(self):
        super().__enter__()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.save_file()
        super().__exit__(exc_type, exc_val, exc_tb)


    def get_sheet(self, sheet_name: str):
        """удаляет лист в книге и создает новый"""
        if sheet_name in (x.title for x in self.workbook.worksheets):
            self.workbook.remove(self.workbook[sheet_name])
        self.sheet = self.workbook.create_sheet(sheet_name)
        self.sheet.font = self.fonts["default"]
        if sheet_name in self.colors.keys():
            self.sheet.sheet_properties.tabColor = self.colors[sheet_name]
        else:
            self.sheet.sheet_properties.tabColor = self.colors["default"]
        return self.sheet


    def create_sheets(self):
        """удаляет все листы в книге и создает новые"""
        if self.workbook:
            for sheet in self.workbook.worksheets:
                self.workbook.remove(sheet)
            # DEFAULT_FONT = self.fonts["default"]
            for i, name in enumerate(self.sheet_names):
                self.sheet = self.workbook.create_sheet(name)
                self.sheet.font = self.fonts["default"]
                self.sheet.sheet_properties.tabColor = self.colors[name]


    def write_header(self, sheet_name: str, header: list, row: int = 1):
        """записывает заголовок в лист"""
        # self.create_sheets()
        self.worksheet = self.workbook[sheet_name]
        # self.worksheet.append(header)

        for col in range(1, len(header) + 1):
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = header[col - 1]
            cell.font = self.fonts["default"]
            cell.fill = self.fills["header"]
            cell.alignment = cell.alignment.copy(wrap_text=True)

    def write_row(self, sheet_name: str, values: list, row_index: int = 2):
        """записывает строку в лист"""
        worksheet = self.workbook[sheet_name]
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index).value = value
            # worksheet.cell(row=row_index, column=column_index).font = self.font

    def write_format(self, sheet_name: str, row_index: int, len_row):
        """записывает строку в лист"""

        self.worksheet = self.workbook[sheet_name]
        for column_index in range(14, 18):
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["default"]
            cell.number_format = self.number_format

        self.worksheet.cell(row=row_index, column=1).font = self.fonts["default_bold"]

        price_cells = [2, 10, 12]
        for column_index in price_cells:
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["green_bold"]
            cell.number_format = self.number_format

        for col in range(3, 9):
            cell = self.worksheet.cell(row=row_index, column=col)
            cell.font = self.fonts["grey"]

        index_cells = (9, 11)
        for column_index in index_cells:
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["blue"]

    def write_material_format(self, sheet_name: str, row_index: int, history_len: int):
        """"""
        columns = {
            "row_number": 1,
            "code": 2,
            "title": 3,
            }

        row_number_cell = self.worksheet.cell(row=row_index, column=columns["row_number"])
        row_number_cell.font = self.fonts["grey"]

        code_cell = self.worksheet.cell(row=row_index, column=columns["code"])
        code_cell.font = self.fonts["default_bold"]

        title_cell = self.worksheet.cell(row=row_index, column=columns["title"])
        title_cell.font = self.fonts["default"]

        price_cells = [
            columns["title"] + 1,
            columns["title"] + 3 + history_len,
            columns["title"] + 5 + history_len,
        ]
        for column_index in price_cells:
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["green_bold"]
            cell.number_format = self.number_format
        #  history
        start_col = columns["title"] + 2
        for col in range(start_col, start_col + history_len):
            cell = self.worksheet.cell(row=row_index, column=col)
            cell.font = self.fonts["grey"]
            cell.number_format = self.number_format
        # text
        text_cells = [
            columns["title"] + history_len + 2,
            columns["title"] + history_len + 4,
            columns["title"] + history_len + 6,
            columns["title"] + history_len + 7,
        ]
        for column_index in text_cells:
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["blue"]


        # transport price
        transport_cells = [
            columns["title"] + history_len + 8,
            columns["title"] + history_len + 9,
            columns["title"] + history_len + 10,
            columns["title"] + history_len + 11,
        ]
        for column_index in transport_cells:
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["default"]
            cell.number_format = self.number_format
        # формулы
        start_col = columns["title"] + history_len + 13
        for column_index in range(start_col, start_col + 11):
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["default"]
            cell.number_format = self.number_format
            cell.fill = self.fills["calculate"]
        # result
        result_col = columns["title"] + history_len + 14
        cell_result = self.worksheet.cell(row=row_index, column=result_col)
        cell_result.font = self.fonts["result_bold"]
        cell_result.number_format = self.number_format
        # percentage of change
        percentage_cols = [
            columns["title"] + history_len + 18,
            columns["title"] + history_len + 23,
        ]
        for column in percentage_cols:
            self.worksheet.cell(row=row_index, column=column).number_format = "0.00%"

        # dangerous flag
        flag_col = columns["title"] + history_len + 22
        flag_cell = self.worksheet.cell(row=row_index, column=flag_col)
        flag_cell.font = self.fonts["result_bold"]
        flag_cell.number_format = "# ##0"

        # width
        slime_cells = [
            1,
            columns["title"] + history_len + 12,
            columns["title"] + history_len + 19,
            columns["title"] + history_len + 22,
        ]
        for column in slime_cells:
            self.worksheet.column_dimensions[get_column_letter(column)].width = 3


