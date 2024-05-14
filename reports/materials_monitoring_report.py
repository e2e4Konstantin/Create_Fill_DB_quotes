import sqlite3
from icecream import ic
import numpy as np
from openpyxl.utils import get_column_letter
from typing import Optional

from config import dbTolls, LocalData
from reports.sql_materials_report import sql_materials_reports
from collections import namedtuple
from dataclasses import dataclass, field

from reports.report_excel_config import ExcelReport
from files_features import output_message_exit


PriceHistory = namedtuple(
    typename="PriceHistory", field_names=["history_index", "history_price"]
)
PriceHistory.__annotations__ = {"history_index": int, "history_price": float}

def sum_of_differences(values):
    if values is None or len(values) <= 3:
        return 1
    mean = sum(values) / len(values)
    return sum(abs(value - mean) for value in values)


def abbe_criterion(signal):
    """Критерий Аббе."""
    if signal is None or len(signal) <= 3:
        return 0
    differences = np.diff(signal) ** 2
    squares = np.sum(differences)
    return np.sqrt(squares / (len(signal) - 1))



@dataclass
class Material:
    code: str
    name: str
    base_price: float
    index_num: int
    actual_price: float
    monitoring_index: int
    monitoring_price: float
    transport_flag: bool = False
    transport_code: str = ""
    transport_base_price: float = 0.0
    transport_factor: float = 0.0
    gross_weight: float = 0.0
    history: list[PriceHistory] = field(default_factory=list)
    len_history: int = 0
    abbe_criterion: float = 0.0
    mean_history: float = 0.0
    std_history: float = 0.0

    def __post_init__(self):
        if self.history is None:
            raise ValueError("history is None")
        self.len_history = len(self.history)
        self.abbe_criterion = abbe_criterion([x.history_price for x in self.history])
        self.mean_history = np.mean([x.history_price for x in self.history])
        self.std_history = np.std([x.history_price for x in self.history])



def _get_material_price_history(
    db: dbTolls, transport_cost_id: int, history_depth: int
) -> list[PriceHistory]:
    """Получить историю цен на материал по id."""
    history = db.go_select(
        sql_materials_reports[
            "select_historical_prices_for_materials_id_not_empty_actual_price"
        ],
        (transport_cost_id, history_depth),
    )
    result = [
        PriceHistory(history_index=row["index_num"], history_price=row["actual_price"])
        for row in history
    ]
    result.sort(reverse=False, key=lambda x: x.history_index)
    return result


def _materials_constructor(
    db: dbTolls, row: sqlite3.Row, history_depth: int
) -> Material:
    material = Material(
        code=row["code"],
        name=row["title"],
        base_price=row["base_price"],
        index_num=row["index_num"],
        actual_price=row["actual_price"],
        monitoring_index=row["monitoring_index_num"],
        monitoring_price=row["monitoring_price"],
        transport_flag=bool(row["transport_flag"]),
        transport_code=row["transport_code"],
        transport_base_price=row["transport_base_price"],
        transport_factor=row["transport_factor"],
        gross_weight=row["gross_weight"],
        history=_get_material_price_history(db, row["ID_tblMaterial"], history_depth),
    )

    return material


def get_materials_with_monitoring( db_file: str, history_depth: int) -> list[Material] | None:
    """Стоимость материалов с данными последнего загруженного мониторинга."""
    table = None
    with dbTolls(db_file) as db:
        # получить id записей материалов с максимальным индексом периода и данными мониторинга
        materials = db.go_select(
            sql_materials_reports["select_records_for_max_index_with_monitoring"]
        )
        table = [_materials_constructor(db, line, history_depth) for line in materials]
    return table if table else None


def fetch_materials_history(db: dbTolls, index_number: int) -> list[sqlite3.Row] | None:
    """Получение истории материалов для заданного номера индекса."""
    materials = db.go_select(
        sql_materials_reports["select_history_material_for_period"],
        {"index_number": index_number},
    )
    return materials


def _fetch_products_history(db: dbTolls, product_id: int, supplement_number: int) -> sqlite3.Row | None:
    """Получение данных продукта из истории по id для заданного номера дополнения."""
    product = db.go_select(
                sql_materials_reports["select_history_product_for_period"],
                {"rowid": product_id, "supplement_number": supplement_number}
       )
    return product[0] if product else None

# select_history_transport_cost_for_perio
def _fetch_transport_cost_history(
    db: dbTolls, transport_cost_id: int, index_number: int
) -> sqlite3.Row | None:
    """Получение данных транспортных расходов из истории по id для заданного номера индекса."""
    transport_cost = db.go_select(
        sql_materials_reports["select_history_transport_cost_for_period"],
        {"rowid": transport_cost_id, "index_number": index_number},
    )
    return transport_cost[0] if transport_cost else None


def _get_materials_for_period_from_history(
    db_file: str, index_number: int
) -> list[Material] | None:
    """Стоимость материалов для периода."""
    table = []
    with dbTolls(db_file) as db:
        materials = fetch_materials_history(db, index_number)
        for line in materials:
            product_id = line["FK_tblMaterials_tblProducts"]
            supplement_number = line["index_supplement_number"]
            product = _fetch_products_history(db, product_id, supplement_number)
            if not product:
                output_message_exit(
                    f"не найден продукт для материала: {line['material_rowid']=}",
                    f"индекс: {index_number} дополнение: {supplement_number}",
                )
            transport_cost_id = line["FK_tblMaterials_tblTransportCosts"]
            transport = _fetch_transport_cost_history(db, transport_cost_id, index_number)
            if not product:
                output_message_exit(
                    f"не найдены транспортные расходы для материала: {line['material_rowid']=}",
                    f"индекс: {index_number} дополнение: {supplement_number}",
                )
            result = dict(product) | dict(line) | dict(transport)
            table.append(result)

    return table if table else None


def _header_create(table: list[Material], view_history_depth: int) -> str:
    header = [
            "No",
            "шифр",
            "название",
            "базовая стоимость",
            "actual index",
            "actual price",
            "monitoring index",
            "monitoring price",
            "transport flag",
            "transport code",
            "transport base price",
            "transport factor",
            "gross weight",
        ]
    header_calculated = [
            "_",
            "стоимость перевозки",
            "цена для загрузки",
            "предыдущий индекс",
            "текущий индекс",
            "рост абс.",
            "рост %",
            "_",
            "abbe критерий",
            'абс.', 'флаг',
            'процент',
        ]

    history_sizes = set([material.len_history for material in table])
    ic(history_sizes)  #  {1, 4, 5}
    max_history_len = max(history_sizes)
    history_header = []
    #  найти материал с длинной историей и записать заголовок
    for material in table:
        if material.len_history == max_history_len:
            # ic(material)
            history_header = [x.history_index for x in material.history]
            #
            if max_history_len > view_history_depth:
                history_header = history_header[-view_history_depth:]
                max_history_len = len(history_header)
            break
    final_header = [*header[:4], *history_header, *header[4:], *header_calculated]
    return final_header, max_history_len


def _material_row_create(material: Material, row_number: int, max_history_len: int):
    base_value = [
        0,
        material.code,
        material.name,
        material.base_price,
        material.index_num,
        material.actual_price,
        material.monitoring_index,
        material.monitoring_price,
        material.transport_flag,
        material.transport_code,
        material.transport_base_price,
        material.transport_factor,
        material.gross_weight,
        "",
    ]
    history_value = [x.history_price for x in material.history]
    if len(history_value) < max_history_len:
        for _ in range(max_history_len - len(history_value)):
            history_value.insert(0, None)
    else:
        history_value = history_value[-max_history_len:]
    #
    start_formula_column = len(base_value) + max_history_len

    cols = {
        "base_price": get_column_letter(start_formula_column - max_history_len - 10),
        "actual_price": get_column_letter(start_formula_column - 8),
        "monitoring_price": get_column_letter(start_formula_column - 6),
        "transport_flag": get_column_letter(start_formula_column - 5),
        "transport_base_price": get_column_letter(start_formula_column - 3),
        "transport_factor": get_column_letter(start_formula_column - 2),
        "gross_weight": get_column_letter(start_formula_column - 1),
        "transport_price": get_column_letter(start_formula_column + 1),
        "result_price": get_column_letter(start_formula_column + 2),
        "previous_index": get_column_letter(start_formula_column + 3),
        "result_index": get_column_letter(start_formula_column + 4),
        "abbe_criterion": get_column_letter(start_formula_column + 8),
        "absolute_price_change": get_column_letter(start_formula_column + 9),
    }
    # ic(cols["transport_flag"])
    formulas = [
        f"={cols['transport_base_price']}{row_number}*{cols['transport_factor']}{row_number}*{cols['gross_weight']}{row_number}/1000",
        f"=ROUND(IF({cols['transport_flag']}{row_number}, ({cols['monitoring_price']}{row_number}-{cols['transport_price']}{row_number}), {cols['monitoring_price']}{row_number}),2)",
        f"={cols['actual_price']}{row_number}/{cols['base_price']}{row_number}",
        f"={cols['result_price']}{row_number}/{cols['base_price']}{row_number}",
        f"={cols['result_index']}{row_number}-{cols['previous_index']}{row_number}",
        f"=({cols['result_index']}{row_number}/{cols['previous_index']}{row_number})-1",
        "",
        round(material.abbe_criterion, 3),
        f"=ROUND(ABS({cols['result_price']}{row_number}-{cols['actual_price']}{row_number}),3)",
        f'=IF({cols["absolute_price_change"]}{row_number}<=1.5*{cols["abbe_criterion"]}{row_number}, "", 1)',
        f"=({cols['result_price']}{row_number}/{cols['actual_price']}{row_number})-1",
    ]
    mod_row = [*base_value[:4], *history_value, *base_value[4:], *formulas]
    # mod_row.append(abbe)
    return mod_row

def materials_monitoring_report_output(table):
    """Напечатать отчет по мониторингу материалов"""
    sheet_name = "materials"
    file_name = "report_monitoring.xlsx"
    with ExcelReport(file_name) as file:
        sheet = file.get_sheet(sheet_name)
        ic(sheet)
        #
        header, max_history_len = _header_create(table)
        file.write_header(sheet.title, header)
        #
        row = 2
        for i, material in enumerate(table):
            value_row = _material_row_create(material, row, max_history_len)
            value_row[0] = i + 1
            file.write_row(sheet_name, value_row, row)
            file.write_material_format(sheet_name, row, max_history_len)
            row += 1
        ic()


def modern_materials_monitoring_report_output(
    table: list[Material], view_history_depth: int, sheet_name: str, file_name: str
) -> None:
    """Напечатать отчет по мониторингу материалов"""
    with ExcelReport(file_name) as file:
        sheet = file.get_sheet(sheet_name)
        ic(sheet)
        #
        header, max_history_len = _header_create(table, view_history_depth)
        file.write_header(sheet.title, header)
        #
        row = 2
        for i, material in enumerate(table):
            value_row = _material_row_create(material, row, max_history_len)
            value_row[0] = i + 1
            file.write_row(sheet_name, value_row, row)
            file.write_material_format(sheet_name, row, max_history_len)
            row += 1
        ic()


def _create_history_price_row(material: Material, max_history_len: int) -> list[Optional[float]]:
    """Создать строку с историей цен на материал"""
    first = [None] * (max_history_len - len(material.history))
    res = (
        first
        + [float(data.history_price) for data in material.history]
        + [
            round(material.abbe_criterion, 3)
            if material.abbe_criterion is not None
            else None
        ]
        + [
            round(material.mean_history, 3)
            if material.mean_history is not None
            else None
        ]
        + [
            round(material.std_history, 3)
            if material.std_history is not None
            else None
        ]
    )


    return res

def _material_price_history_report_output(table, sheet_name: str, file_name: str) -> list[str] | None:
    """Напечатать отчет по ценам материалов"""
    max_history_len = max(material.len_history for material in table)
    for material in table:
        if material.len_history == max_history_len:
            history_header = [str(x.history_index) for x in material.history]
            break
    history_header = ["No", "code", *history_header, "abbe criterion", "mean", "std"]

    row = 2
    with ExcelReport(file_name) as file:
        sheet = file.get_sheet(sheet_name)
        if history_header:
            sheet.append(history_header)
        for i, material in enumerate(table):
            price_row = _create_history_price_row(material, max_history_len)
            price_row = [i + 1, material.code, *price_row]
            sheet.append(price_row)
            row += 1




if __name__ == "__main__":


    location = "office"  # office  # home

    local = LocalData(location)
    ic()

    table = _get_materials_for_period_from_history(
        local.db_file, index_number=211)
    ic(len(table))
    ic(table[5])

    # table = get_materials_with_monitoring(local.db_file, history_depth=10)

    # # for material in table:
    # #     if material.len_history == 7:
    # #         ic(material.code)

    # ic(len(table))
    # ic( table[3])
    # # check_list = [
    # #     "1.1-1-8106",
    # #     "1.1-1-8107",
    # #     "1.1-1-1",
    # #     "1.1-1-8108",
    # #     "1.21-5-2107",
    # #     "1.21-5-2116",
    # # ]
    # # check_table = []
    # # for material in table:
    # #     if material.code in check_list:
    # #         ic(material)
    # #         check_table.append(material)
    # sheet_name = "materials"
    # file_name = "report_monitoring.xlsx"
    # modern_materials_monitoring_report_output(
    #     table, view_history_depth=2, sheet_name =sheet_name, file_name = file_name
    # )
    # _material_price_history_report_output(
    #     table, sheet_name="price materials history", file_name=file_name
    # )
