import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
import os
from files_features import output_message_exit


class ExcelControl:
    def __init__( self, excel_file: str = None ):
        self.file = excel_file
        self.book = None
        self.sheet = None
        self.sheet_names = ["транспорт", "мониторинг",]
        self.sheet_colors = [
            "00FF9900",
            "0099CC00",
            "00FFCC00",
            "000066CC",
            "00666699",
            "00C0C0C0",
            "00FF99CC",
        ]

    def __enter__(self):
        self.open_file()
        self.create_sheets()
        return self

    def __exit__(self, exception_type, exception_value, traceback):
        self.close_file()

    def __str__(self):
        return (
            f"excel file: {self.file_name}, sheet: {self.sheet.sheets} "
        )
    def file_in_use(self) -> bool:
        """файл используется другой программой."""
        filepath = self.file
        return os.path.isfile(filepath) and not os.access(filepath, os.W_OK)


    def open_file(self):
        """Открывает файл, если он есть или создает его"""
        if self.file_in_use():
            output_message_exit(f"файла: {self.file}", "занят другой программой")
        try:
            self.book = openpyxl.load_workbook(self.file)

        except IOError as err:
            print(f"ошибка при открытии файла: {self.file}\n{err}")
            raise

    def close_file(self):
        if self.book:
            self.book.close()

    def save_file(self):
        """save the workbook"""
        if self.file_in_use():
            self.close_file()
            output_message_exit(
                "Не могу записать файл",
                f"{self.file} используется в другой программе.",
            )
        else:
            if self.book:
                self.book.save(self.file)

    def sheets_create(self):
        """Удаляет все листы в книге и создает новые."""

        if self.book:
            for sheet in self.book.worksheets:
                self.book.remove(sheet)
            for name in self.sheet_names:
                sheet = self.book.create_sheet(name)
                sheet.sheet_properties.tabColor = self.tab_color[name]

    def header_write(self, sheet, header_list):
        if sheet and len(header_list) > 0:
            sheet.append(header_list)
            for cell in range(1, len(header_list) + 1):
                sheet.cell(row=1, column=cell).fill = PatternFill(
                    "solid", fgColor=self.tab_color["Header"]
                )

    def save_quotes(self, quotes_data):
        sheet_name = self.items_quote_set[0]  # 'Quote'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "GROUP_WORK_PROCESS",
                "PRESSMARK",
                "TITLE",
                "UNIT_OF_MEASURE",
                "STAT_SUM",
                "PARAMETERIZED_FLAG",
                "SUPPLEMENTARY_TYPE",
                "PARENT_PRESSMARK",
                "ALGORITHM",
            ]
            self.header_write(sheet, header)
            for quote in quotes_data:
                line_data = [getattr(quote, x.name) for x in fields(Quote)]
                line_data[6] = "++" if line_data[6] else " "
                line_data[9] = line_data[9] if line_data[9] > 0 else " "
                sheet.append(line_data[:-2])
                # sheet.append(line_data[:-2])
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 100
            sheet.column_dimensions["G"].width = 20
        else:
            print(
                f"save_quote >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_attributes(self, quotes_data):
        sheet_name = self.items_quote_set[1]  # 'Attributes'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "PRESSMARK",
                "ATTRIBUTE_TITLE",
                "VALUE",
            ]
            self.header_write(sheet, header)
            data_line = list()
            for quote in quotes_data:
                for attribute in quote.attributes_quote:
                    data_line.clear()
                    data_line.append(quote.row_quote)
                    data_line.append(quote.cod_quote)
                    data_line.append(
                        string_cleaning_capitalize(attribute.name_attribute)
                    )
                    data_line.append(
                        string_cleaning_capitalize(attribute.value_attribute)
                    )
                    sheet.append(data_line)
            sheet.column_dimensions["B"].width = 12
            sheet.column_dimensions["C"].width = 40
            sheet.column_dimensions["D"].width = 60
        else:
            print(
                f"save_attributes >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_options(self, quotes_data):
        sheet_name = self.items_quote_set[2]  # 'Options'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "PRESSMARK",
                "PARAMETER_TITLE",
                "LEFT_BORDER",
                "RIGHT_BORDER",
                "UNIT_OF_MEASURE",
                "STEP",
                "PARAMETER_TYPE",
            ]
            self.header_write(sheet, header)
            data_line = list()
            for quote in quotes_data:
                for option in quote.options_quote:
                    data_line.clear()
                    data_line.append(quote.row_quote)
                    data_line.append(quote.cod_quote)
                    data_line.append(string_cleaning_capitalize(option.name_option))
                    for value in option.value_option:
                        data_line.append(value[1])
                    sheet.append(data_line)
            sheet.column_dimensions["C"].width = 70
        else:
            print(
                f"save_options >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_collections(self, collections_data):
        sheet_name = self.items_quote_set[3]  # 'Collections'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]

            header = [
                "Строка",
                "Шифр",
                "Наименование сборника",
                "Всего расценок",
                "Параметризованных",
                "Пустых",
                "пустые расценки",
            ]
            self.header_write(sheet, header)

            for collection in collections_data:
                line_data = [
                    getattr(collections_data[collection], x.name)
                    for x in fields(Collection)
                ]
                total_quotes = (
                    collections_data[collection].quantity_parameterized_quotes
                    + collections_data[collection].quantity_not_parameterized_quotes
                )
                line_data.insert(3, total_quotes)
                bad_list = line_data.pop()
                bad_list_string = ", ".join(x for x in bad_list)
                line_data.append(bad_list_string)
                sheet.append(line_data)
            sheet.column_dimensions["C"].width = 100
        else:
            print(
                f"save_collections >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_tables(self, tables_data):
        sheet_name = self.items_quote_set[4]  # 'Tables'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "row",
                "номер",
                "код",
                "атрибутов",
                "параметров",
                "название",
                "атрибуты",
                "параметры",
            ]
            self.header_write(sheet, header)
            for table in tables_data:
                data_line = tables_data[table].to_list()
                sheet.append(data_line)
        else:
            print(
                f"save_tables >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_console(self, text_in: str):  # 'Console' 5
        sheet_name = self.items_quote_set[5]
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            # for x in range(1, 11):
            #     sheet.row_dimensions[x].height = 140
            # sheet.column_dimensions['A'].width = 50
            # sheet.column_dimensions['B'].width = 50
            # sheet['A1'] = text_inget_column_number
            # sheet.merge_cells('A1:B10')
            # sheet['A1'].font = Font(size=9)
            # sheet['A1'].alignment = Alignment(horizontal='left', vertical="top", wrapText=True) #"bottom"

            sheet.column_dimensions["A"].width = 100
            text_by_lines = text_in.splitlines()
            for i, line in enumerate(text_by_lines):
                sheet.append((line,))
                sheet.cell(row=i + 1, column=1).font = Font(size=10)
                sheet.cell(row=i + 1, column=1).alignment = Alignment(
                    horizontal="left", vertical="bottom", wrapText=True
                )
        else:
            print(
                f"save_console >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_failed_tables(
        self, broken_tables: list[tuple[int, str]]
    ):  # 'BrokenTable' 6
        sheet_name = self.items_quote_set[6]  # 'BrokenTable'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = ["row", "название таблицы"]
            self.header_write(sheet, header)
            sheet = self.book[sheet_name]
            for table in broken_tables:
                sheet.append([table[0], table[1]])
        else:
            print(
                f"save_failed_tables >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_resources(self, resource_data):
        sheet_name = self.items_resources_set[0]  # Resources
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "A",
                "B",
                "PRESSMARK",
                "TITLE",
                "UOM",
                "OKP",
                "USE_COUNT",
                "PARAMETERIZED_FLAG",
                "TABLE",
            ]  #
            self.header_write(sheet, header)
            # data_line = list()
            for resource in resource_data:
                line_data = [getattr(resource, x.name) for x in fields(Resource)]
                line_data[8] = "++" if line_data[7] else " "
                line_data[9] = resource_tables[resource.table].cod_table
                sheet.append(line_data[:-2])
            sheet.column_dimensions["D"].width = 15
            sheet.column_dimensions["E"].width = 80
            sheet.column_dimensions["F"].width = 20
            sheet.column_dimensions["G"].width = 10
            sheet.column_dimensions["H"].width = 20
            for row in sheet[2 : sheet.max_row]:  # пропускаем заголовок
                cell_f = row[5]  # column F
                cell_f.alignment = Alignment(horizontal="center")
                cell_f = row[7]  # column H
                cell_f.alignment = Alignment(horizontal="center")
        else:
            print(
                f"save_resources >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    #
    def save_resources_attributes(self, resource_data):
        sheet_name = self.items_resources_set[1]  # 'Attributes'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "PRESSMARK",
                "ATTRIBUTE_TITLE",
                "VALUE",
            ]
            self.header_write(sheet, header)
            data_line = list()
            for resource in resource_data:
                for attribute in resource.attributes_resource:
                    data_line.clear()
                    data_line.append(resource.row)
                    data_line.append(resource.press_mark)
                    data_line.append(attribute.name_attribute)
                    data_line.append(attribute.value_attribute)
                    sheet.append(data_line)
            sheet.column_dimensions["B"].width = 13
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 50
        else:
            print(
                f"save_resources_attributes >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_resources_options(self, resource_data):
        sheet_name = self.items_resources_set[2]  # 'Options'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "PRESSMARK",
                "PARAMETER_TITLE",
                "LEFT_BORDER",
                "RIGHT_BORDER",
                "UNIT_OF_MEASURE",
                "STEP",
                "PARAMETER_TYPE",
            ]
            self.header_write(sheet, header)
            data_line = list()
            for resource in resource_data:
                for option in resource.options_resource:
                    data_line.clear()
                    data_line.append(resource.row)
                    data_line.append(resource.press_mark)
                    data_line.append(option.name_option)
                    for value in option.value_option:
                        data_line.append(value[1])
                    sheet.append(data_line)
            sheet.column_dimensions["B"].width = 12
            sheet.column_dimensions["C"].width = 25
        else:
            print(
                f"save_resources_options >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_equipment(self, equipment_data):
        sheet_name = self.items_resources_set[0]  # Equipment
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "A",
                "B",
                "PRESSMARK",
                "TITLE",
                "UOM",
                "USE_COUNT",
                "PARAMETERIZED_FLAG",
                "REMARK",
                "TABLE",
            ]
            self.header_write(sheet, header)
            # line_data = list()
            for equipment in equipment_data:
                line_data = [getattr(equipment, x.name) for x in fields(Equipment)]
                line_data[7] = "++" if line_data[7] else " "
                line_data[9] = equipment_tables[equipment.table].cod_table
                sheet.append(line_data[:-2])

            sheet.column_dimensions["D"].width = 15
            sheet.column_dimensions["E"].width = 80
            sheet.column_dimensions["F"].width = 20
            sheet.column_dimensions["G"].width = 10
            sheet.column_dimensions["H"].width = 20
            for row in sheet[2 : sheet.max_row]:  # пропускаем заголовок
                cell_f = row[5]  # column F
                cell_f.alignment = Alignment(horizontal="center")
                cell_f = row[7]  # column H
                cell_f.alignment = Alignment(horizontal="center")
        else:
            print(
                f"save_equipment >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_equipment_attributes(self, equipment_data):
        sheet_name = self.items_resources_set[1]  # 'Attribute'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "PRESSMARK",
                "ATTRIBUTE_TITLE",
                "VALUE",
            ]
            self.header_write(sheet, header)
            data_line = list()
            for equipment in equipment_data:
                for attribute in equipment.attributes_equipment:
                    data_line.clear()
                    data_line.append(equipment.row)
                    data_line.append(equipment.press_mark)
                    data_line.append(attribute.name_attribute)
                    data_line.append(attribute.value_attribute)
                    sheet.append(data_line)
            sheet.column_dimensions["B"].width = 13
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 50
        else:
            print(
                f"save_equipment_attributes >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_equipment_options(self, equipment_data):
        sheet_name = self.items_resources_set[2]  # 'Options'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "ROW",
                "PRESSMARK",
                "PARAMETER_TITLE",
                "LEFT_BORDER",
                "RIGHT_BORDER",
                "UNIT_OF_MEASURE",
                "STEP",
                "PARAMETER_TYPE",
            ]
            self.header_write(sheet, header)
            data_line = list()
            for equipment in equipment_data:
                for option in equipment.options_equipment:
                    data_line.clear()
                    data_line.append(equipment.row)
                    data_line.append(equipment.press_mark)
                    data_line.append(option.name_option)
                    for value in option.value_option:
                        data_line.append(value[1])
                    sheet.append(data_line)
            sheet.column_dimensions["B"].width = 12
            sheet.column_dimensions["C"].width = 25
        else:
            print(
                f"save_equipment_options >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_equipment_tables(self, tables_data):
        sheet_name = self.items_resources_set[4]  # 'Tables'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "row",
                "номер",
                "код",
                "атрибутов",
                "параметров",
                "название",
                "атрибуты",
                "параметры",
            ]
            self.header_write(sheet, header)
            for table in tables_data:
                data_line = table.to_list()
                sheet.append(data_line)
        else:
            print(
                f"save_equipment_tables >> в файле {self.full_path} не найден лист {sheet_name}."
            )

    def save_resources_tables(self, tables_data):
        sheet_name = self.items_resources_set[4]  # 'Tables'
        if sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            header = [
                "row",
                "номер",
                "код",
                "атрибутов",
                "параметров",
                "название",
                "атрибуты",
                "параметры",
            ]
            self.header_write(sheet, header)
            for table in tables_data:
                data_line = table.to_list()
                sheet.append(data_line)
        else:
            print(
                f"save_resources_tables >> в файле {self.full_path} не найден лист {sheet_name}."
            )
