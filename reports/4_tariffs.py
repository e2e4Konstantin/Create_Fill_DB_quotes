import sqlite3
from icecream import ic
from config import ExcelBase
from reports.report_excel_config import ExcelReport
from config import dbTolls, LocalData, ROUNDING
from numbers import Number
from openpyxl.worksheet.worksheet import Worksheet
from reports.sql_materials_report import sql_materials_reports


def _find_tariff_code(ws: Worksheet, row: int, max_cols: int) -> tuple[int, str] | None:
    """Найти код тарифа в строке листа."""
    codes = ("1.1-1-118", "1.1-1-41", "1.1-1-1268", "1.1-1-2238", "1.1-1-1920")
    for col in range(1, max_cols + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            val = str(val).strip()
            if val in codes:
                return col, val
    return None


def _get_last_number_in_row(
    worksheet: Worksheet, row: int, max_columns: int, start_column: int = 1
) -> tuple[int, float] | None:
    """
    Находит последнее число в строке рабочего листа Excel.

    Args:
        worksheet: Рабочий лист.
        row: номер строки для поиска.
        max_columns: Максимальный индекс столбца.
        start_column: Индекс первого столбца для поиска (по умолчанию 1).

    Returns:
        AКортеж, содержащий индекс столбца и значение последнего найденного числа, или
        None, если не найдено ни одного числа.
    """
    last_number = None
    for column in range(start_column, max_columns + 1):
        value = worksheet.cell(row=row, column=column).value
        if isinstance(value, (int, float)):
            last_number = (column, float(value))
    return last_number



def _tariff_file_parsing(tariffs_file_name: str, sheet_name: str = "Тарифы") -> tuple[tuple[str, float], ...] | None:
    """
    Читает файл с тарифами Excel, находит коды тарифов и последние значения цен
    """
    tariffs = []
    with ExcelBase(tariffs_file_name) as file:
        worksheet = file.workbook[sheet_name]
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        for row in range(1, max_row + 1):
            code_cell = _find_tariff_code(worksheet, row, max_column)
            if code_cell:
                last_number = _get_last_number_in_row(worksheet, row, max_column, start_column=code_cell[0])
                if last_number is not None:
                    tariffs.append((code_cell[1], last_number[1]))
                else:
                    tariffs.append((code_cell[1], 0.0))
    return tuple(tariffs)


def _get_data_material(
    db: dbTolls, supplement_num: int, code_material: str
) -> None:
    """ Получает информацию о материале из истории по шифру для номера дополнения"""
    result = db.go_select(
        sql_materials_reports["select_product_from_history_for_supplement_by_code"],
        ({"supplement": supplement_num, "code": code_material}),
    )
    if result:
        _rowid, period_title, code, description, measurer = result[0]
        return period_title, description, measurer
    return None


def tariffs_monitoring_report(
    tariffs_data, db_name: str, sheet_name: str, file_name: str):
    header = ["период","No", "шифр", "описание", "ед.изм", "цена текущая"]
    with ExcelReport(file_name) as report_file:
        sheet = report_file.get_sheet(sheet_name)
        sheet.append(header)
        report_file.set_tariffs_header_format(sheet_name, row=1, header_len=len(header))
        for num, tariff in enumerate(tariffs_data, start=1):
            period_title, code, description, measurer, price = tariff
            line = [period_title, num, code, description, measurer, price]
            sheet.append(line)
            report_file.set_regular_row_tariff_format(sheet_name, row=num + 1)


def _update_tariffs_from_db(tariffs: tuple[tuple[str, float], ...], db_file: str, supplement_num: int) -> None:
    with dbTolls(db_file) as db:
        new_tariffs = []
        for tariff in tariffs:
            code, price = tariff
            data = _get_data_material(db, supplement_num, code)
            period_title, description, measurer = data
            if data:
                new_tariffs.append((period_title, code, description, measurer, price))
        return tuple(new_tariffs) if new_tariffs else None

if __name__ == "__main__":
    location = "office"  # office  # home
    local = LocalData(location)
    ic()
    report_file_name = "report_monitoring.xlsx"
    tariffs_file_name = r"C:\Users\kazak.ke\Documents\Задачи\5_Надя\исходные_данные\тарифы\Тарифы на 20.05.2024.xlsx"

    tariffs = _tariff_file_parsing(tariffs_file_name=tariffs_file_name, sheet_name="Тарифы")
    ic(len(tariffs))
    ic(tariffs)
    tariffs = _update_tariffs_from_db(tariffs, local.db_file, 72)
    ic(tariffs)

    tariffs_monitoring_report(
        db_name=local.db_file,
        tariffs_data=tariffs,
        sheet_name="tariff_prices",
        file_name=report_file_name,
    )
