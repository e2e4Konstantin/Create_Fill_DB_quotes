import openpyxl
from openpyxl.styles import numbers


def write_formulas(workbook_path, sheet_name, formulas):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[sheet_name]

    for row, col, formula in formulas:
        worksheet.cell(row=row, column=col, value=formula)

    a1 = worksheet.cell(row=1, column=1, value=4587878.123)
    a1.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    a2 = worksheet.cell(row=1, column=2, value=4587878.123)
    a2.number_format = "#,##0.00"

    workbook.save(workbook_path)
    workbook.close()


workbook_path = r"C:\Users\kazak.ke\Documents\Tmp\test.xlsx"
sheet_name = "Test 77"
formulas = [
    (5, 1, "=SUM(5, 4)"),
    (2, 2, "=SUM('Test 99'!A1:I1)"),
    (5, 2, "=IF(B1, 55, 33)"),
    (4, 3, "=IF(B1, 55, 33)")
]

write_formulas(workbook_path, sheet_name, formulas)

